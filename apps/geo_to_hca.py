from pysradb.sraweb import SRAweb
from time import sleep
import requests as rq
import pandas as pd
import xml.etree.ElementTree as xm
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl import load_workbook
import os,sys
import re
import argparse
import multiprocessing
from functools import partial
from contextlib import contextmanager
import itertools
from itertools import chain

STATUS_ERROR_CODE = 400

class NotFoundSRA(Exception):
    """
    Sub-class for Exception to handle 400 error status codes.
    """
    def __init__(self, response, accession_list):
        self.response = response
        self.accessions = accession_list
        self.error = self.parse_xml_error()

    def parse_xml_error(self):
        root = xm.fromstring(self.response.content)
        return root.find('ERROR').text  # Return the string for the error returned by Efetch

    def __str__(self):
        if len(self.accessions) > 1:
            accession_string = '\n'.join(self.accessions)
        else:
            accession_string = self.accessions
        return (f"\nStatus code of the request: {self.response.status_code}.\n"
                f"Error as returned by SRA:\n{self.error}"
                f"The provided accessions were:\n{accession_string}\n\n")

class NotFoundENA(Exception):
    """
    Sub-class for Exception to handle 400 error status codes.
    """
    def __init__(self, response, title):
        self.response = response
        self.title = title
        self.error = self.parse_xml_error()

    def parse_xml_error(self):
        root = xm.fromstring(self.response.content)
        return root.find('ERROR').text  # Return the string for the error returned by Efetch

    def __str__(self):
        return (f"\nStatus code of the request: {self.response.status_code}.\n"
                f"Error as returned by ENA:\n{self.error}"
                f"The provided project title or name was:\n{self.title}\n\n")

class SraUtils:

    @staticmethod
    def sra_accession_from_geo(geo_accession: str):
        sleep(0.5)
        try:
            srp = SRAweb().gse_to_srp(geo_accession)
        except:
            srp = None
        if not isinstance(srp, pd.DataFrame):
            srp = None
        elif isinstance(srp, pd.DataFrame):
            if srp.shape[0] == 0:
                srp = None
            else:
                srp = srp
        return srp

    @staticmethod
    def srp_metadata(srp_accession: str) -> pd.DataFrame:
        sleep(0.5)
        srp_metadata_url = f'http://trace.ncbi.nlm.nih.gov/Traces/sra/sra.cgi?save=efetch&db=sra&rettype=runinfo&term={srp_accession}'
        return pd.read_csv(srp_metadata_url)

    @staticmethod
    def get_fastq(srr_accessions):
        sleep(0.5)
        url = f'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch/fcgi?db=sra&id={",".join(srr_accessions)}'
        srr_metadata_url = rq.get(url)
        if srr_metadata_url.status_code == STATUS_ERROR_CODE:
            raise NotFoundSRA(srr_metadata_url, srr_accessions)
        try:
            xml = xm.fromstring(srr_metadata_url.content)
            xml_content = srr_metadata_url.content
        except:
            xml = None
            xml_content = None
        return xml,xml_content,url

    @staticmethod
    def split_list(accessions, n):
        parts_list = []
        for i in range(0, len(accessions), n):
            part = accessions[i:i + n]
            if part:
                parts_list.append(part)
        return parts_list

    @staticmethod
    def request_info(accessions,accession_type):
        if accession_type == 'biosample':
            url = f'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch/fcgi?db=biosample&id={",".join(accessions)}'
        if accession_type == 'experiment':
            url = f'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch/fcgi?db=sra&id={",".join(accessions)}'
        sra_url = rq.get(url)
        if sra_url.status_code == STATUS_ERROR_CODE:
            raise NotFoundSRA(sra_url, accessions)
        return xm.fromstring(sra_url.content)

    @staticmethod
    def get_content(accessions,accession_type):
        if len(accessions) < 100:
            xml = SraUtils.request_info(accessions,accession_type=accession_type)
            size = 'small'
            return xml,size
        else:
            size = 'large'
            parts_list = SraUtils.split_list(accessions, n=100)
            xmls = []
            for p in range(0,len(parts_list)):
                xml = SraUtils.request_info(parts_list[p],accession_type=accession_type)
                xmls.append(xml)
            return xmls,size

    @staticmethod
    def srp_bioproject(bioproject_accession):
        sleep(0.5)
        srp_bioproject_url = rq.get(f'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch/fcgi?db=bioproject&id={bioproject_accession}')
        if srp_bioproject_url.status_code == STATUS_ERROR_CODE:
            raise NotFoundSRA(srp_bioproject_url, bioproject_accession)
        return xm.fromstring(srp_bioproject_url.content)

    @staticmethod
    def pubmed_id(project_pubmed_id):
        sleep(0.5)
        pubmed_url = rq.get(f'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch/fcgi?db=pubmed&id={project_pubmed_id}&rettype=xml')
        if pubmed_url.status_code == STATUS_ERROR_CODE:
            raise NotFoundSRA(pubmed_url, project_pubmed_id)
        return xm.fromstring(pubmed_url.content)

def fetch_srp_accession(geo_accession: str):
    srp = SraUtils.sra_accession_from_geo(geo_accession)
    if not srp.empty:
        if srp.shape[0] == 1:
            srp = srp.iloc[0]["study_accession"]
        elif srp.shape[0] > 1:
            answer = input("More than 1 SRA Study Accession has been found. Is this the intended SRA study id? %s [y/n]: " % (srp.iloc[0]["study_accession"]))
            if answer.lower() in ['y', "yes"]:
                srp = srp.iloc[0]["study_accession"]
            if answer.lower() in ['n', "no"]:
                answer = input("Alternatively, is this the intended SRA study id? %s [y/n]: " % (srp.iloc[1]["study_accession"]))
                if answer.lower() in ['y', "yes"]:
                        srp = srp.iloc[1]["study_accession"]
                if answer.lower() in ['n', "no"]:
                        answer = input("Alternatively, is this the intended SRA study id? %s [y/n]: " % (srp.iloc[2]["study_accession"]))
                        if answer.lower() in ['y', "yes"]:
                            srp = srp.iloc[2]["study_accession"]
                        if answer.lower() in ['n', "no"]:
                            print("SRA study accession not found")
                            srp = None
    else:
        answer = input("Could not recognise GEO accession %s; is it a GEO Superseries? If yes, please enter the project GEO accession manually here (GSExxxxxx) or type exit for program exit: " % (geo_accession))
        srp = SraUtils.sra_accession_from_geo(answer.upper())
        if not srp:
            print("GEO accession still not recognised; exiting program")
        else:
            if srp.shape[0] == 1:
                srp = srp.iloc[0]["study_accession"]
            elif srp.shape[0] > 1:
                answer = input("More than 1 SRA Study Accession has been found. Is this the intended SRA study id? %s [y/n]: " % (srp.iloc[0]["study_accession"]))
                if answer.lower() in ['y', "yes"]:
                    srp = srp.iloc[0]["study_accession"]
                if answer.lower() in ['n', "no"]:
                    answer = input("Alternatively, is this the intended SRA study id? %s [y/n]: " % (srp.iloc[1]["study_accession"]))
                    if answer.lower() in ['y', "yes"]:
                        srp = srp.iloc[1]["study_accession"]
                    if answer.lower() in ['n', "no"]:
                        answer = input("Alternatively, is this the intended SRA study id? %s [y/n]: " % (srp.iloc[2]["study_accession"]))
                        if answer.lower() in ['y', "yes"]:
                            srp = srp.iloc[2]["study_accession"]
                        if answer.lower() in ['n', "no"]:
                            print("SRA study accession not found")
                            srp = None
    return srp

def fetch_srp_metadata(srp_accession: str) -> pd.DataFrame:
    srp_metadata_df = SraUtils.srp_metadata(srp_accession)
    return srp_metadata_df

def get_reads(ftp_path):
    try:
        read_files = ftp_path.split(';')
        read_files = [file.split("/")[-1] for file in read_files]
    except:
        read_files = []
    return read_files

def get_fastq_from_ENA(srp_accession):
    try:
        request_url= f'http://www.ebi.ac.uk/ena/data/warehouse/filereport?accession={srp_accession}&result=read_run&fields=run_accession,fastq_ftp'
        fastq_results = pd.read_csv(request_url, delimiter='\t')
        fastq_map = {list(fastq_results['run_accession'])[i]: get_reads(list(fastq_results['fastq_ftp'])[i]) for i in range(0, len(list(fastq_results['run_accession'])))}
        return fastq_map
    except:
        return None

def extract_read_info(read_values,accession,fastq_map):
    if "--read1PairFiles" not in read_values or "--read2PairFiles" not in read_values:
        return fastq_map
    else:
        read_info_list = read_values.strip().split('--')[1:]
        for read_info in read_info_list:
            read_info = read_info.strip()
            split_read = read_info.split("=")
            if 'fastq.gz' in split_read[1]:
                if accession in fastq_map.keys():
                    fastq_map[accession].append(split_read[1])
                else:
                    fastq_map[accession] = [split_read[1]]
    return fastq_map

def get_fastq_from_SRA(srr_accessions):
    xml_content, content, url = SraUtils.get_fastq(srr_accessions)
    if not xml_content:
        fastq_map = None
    else:
        experiment_packages = parse_xml(xml_content)
        for experiment_package in experiment_packages:
            try:
                fastq_map = get_file_names_from_SRA(experiment_package)
            except:
                continue
    return fastq_map

def test_number_files(fastq_map):
    if fastq_map:
        test_number_files = [len(fastq_map[accession]) < 2 for accession in fastq_map.keys()]
        if all(test_number_files) is True:
            fastq_map = None
        elif any(test_number_files) is True:
            print("Fastq file names for only some of the SRA run accessions are not available.")
    return fastq_map

def fetch_fastq_names(srp_accession, srr_accessions):
    # Try fetching the fstq file names from ENA using the run accession
    fastq_map = get_fastq_from_ENA(srp_accession)
    fastq_map = test_number_files(fastq_map)
    # If fastq files are not available in ENA, try searching in SRA:
    if not fastq_map:
        # First send a request for info. about the list of SRA accessions from SRA
        fastq_map = get_fastq_from_SRA(srr_accessions)
        fastq_map = test_number_files(fastq_map)
    return fastq_map

def fetch_accession_info(accessions_list: [],accession_type):
    xml_content_result,size = SraUtils.get_content(accessions_list,accession_type=accession_type)
    if size == 'large':
        nested_list = []
        if accession_type == 'biosample':
            for xml_content in xml_content_result:
                nested_list.extend([get_attributes_biosample(element) for element in xml_content])
        elif accession_type == 'experiment':
            for xml_content in xml_content_result:
                for experiment_package in xml_content.findall('EXPERIMENT_PACKAGE'):
                    nested_list.extend([get_attributes_library_protocol(experiment_package)])
    else:
        if accession_type == 'biosample':
            nested_list = [get_attributes_biosample(element) for element in xml_content_result]
        elif accession_type == 'experiment':
            nested_list = []
            for experiment_package in xml_content_result.findall('EXPERIMENT_PACKAGE'):
                nested_list.extend([get_attributes_library_protocol(experiment_package)])
    return nested_list

def fetch_bioproject(bioproject_accession: str):
    xml_content = SraUtils.srp_bioproject(bioproject_accession)
    bioproject_metadata = xml_content.find('DocumentSummary')
    try:
        project_name = bioproject_metadata.find("Project").find('ProjectDescr').find('Name').text
    except:
        print("no project name")
        project_name = None
    try:
        project_title = bioproject_metadata.find("Project").find('ProjectDescr').find('Title').text
    except:
        print("no project title")
        project_title = None
    try:
        project_description = bioproject_metadata.find("Project").find('ProjectDescr').find('Description').text
    except:
        project_description = ''
        print("no project description")
    project_publication = bioproject_metadata.find("Project").find('ProjectDescr').find('Publication')
    try:
        if project_publication.find('DbType').text == 'Pubmed' or project_publication.find('DbType').text == 'ePubmed':
            try:
                project_pubmed_id = project_publication.attrib['id']
            except:
                project_pubmed_id = project_publication.find('Reference').text
    except:
        print("No publication for project %s was found: searching project title in EuropePMC" % (bioproject_accession))
    if not project_publication or not project_pubmed_id:
        if project_title:
            print("project title is: %s" % (project_title))
            url = rq.get(f'https://www.ebi.ac.uk/europepmc/webservices/rest/search?query={project_title}')
            if url.status_code == STATUS_ERROR_CODE:
                raise NotFoundENA(url, project_title)
            else:
                xml_content = xm.fromstring(url.content)
                try:
                    results = list()
                    result_list = xml_content.find("resultList")
                    for result in result_list:
                        results.append(result)
                    journal_title = results[0].find("journalTitle").text
                    if not journal_title or journal_title == '':
                        project_pubmed_id = ''
                        print("no publication results for project title in ENA")
                    else:
                        answer = input("A publication title has been found: %s.\nIs this the publication title associated with the GEO accession? [y/n]: " % (journal_title))
                        if answer.lower() in ['y',"yes"]:
                            project_pubmed_id = results[0].find("pmid").text
                        else:
                            journal_title = results[1].find("journalTitle").text
                            if not journal_title or journal_title == '':
                                project_pubmed_id = ''
                                print("no publication results for project title in ENA")
                            else:
                                answer = input("An alternative publication title has been found: %s.\nIs this the publication title associated with the GEO accession? [y/n]: " % (journal_title))
                                if answer.lower() in ['y', "yes"]:
                                    project_pubmed_id = results[1].find("pmid").text
                                else:
                                    journal_title = results[2].find("journalTitle").text
                                    if not journal_title or journal_title == '':
                                        project_pubmed_id = ''
                                        print("no publication results for project title in ENA")
                                    else:
                                        answer = input("An alternative publication title has been found: %s.\nIs this the publication title associated with the GEO accession? [y/n]: " % (journal_title))
                                        if answer.lower() in ['y', "yes"]:
                                            project_pubmed_id = results[2].find("pmid").text
                                        else:
                                            project_pubmed_id = ''
                                            print("no publication results for project title in ENA")
                except:
                    print("no publication results for project title in ENA")
                    project_pubmed_id = ''
        if not project_pubmed_id or project_pubmed_id == '':
            if project_name:
                print("project name is %s:" % (project_name))
                url = rq.get(f'https://www.ebi.ac.uk/europepmc/webservices/rest/search?query={project_name}')
                if url.status_code == STATUS_ERROR_CODE:
                    raise NotFoundENA(url, project_name)
                else:
                    xml_content = xm.fromstring(url.content)
                    try:
                        results = list()
                        result_list = xml_content.find("resultList")
                        for result in result_list:
                            results.append(result)
                        journal_title = results[0].find("journalTitle").text
                        if not journal_title or journal_title == '':
                            project_pubmed_id = ''
                            print("no publication results for project name in ENA")
                        else:
                            answer = input("A publication title has been found: %s.\nIs this the publication title associated with the GEO accession? [y/n]: " % (journal_title))
                            if answer.lower() in ['y',"yes"]:
                                project_pubmed_id = results[0].find("pmid").text
                            else:
                                journal_title = results[1].find("journalTitle").text
                                if not journal_title or journal_title == '':
                                    project_pubmed_id = ''
                                    print("no publication results for project name in ENA")
                                else:
                                    answer = input("An alternative publication title has been found: %s.\nIs this the publication title associated with the GEO accession? [y/n]: " % (journal_title))
                                    if answer.lower() in ['y', "yes"]:
                                        project_pubmed_id = results[1].find("pmid").text
                                    else:
                                        journal_title = results[2].find("journalTitle").text
                                        if not journal_title or journal_title == '':
                                            project_pubmed_id = ''
                                            print("no publication results for project name in ENA")
                                        else:
                                            answer = input("An alternative publication title has been found: %s.\nIs this the publication title associated with the GEO accession? [y/n]: " % (journal_title))
                                            if answer.lower() in ['y', "yes"]:
                                                project_pubmed_id = results[2].find("pmid").text
                                            else:
                                                project_pubmed_id = ''
                                                print("no publication results for project name in ENA")
                    except:
                        print("no publication results for project name in ENA")
                        project_pubmed_id = ''
        if not project_pubmed_id or project_pubmed_id == '':
            project_title = ''
            project_name = ''
            project_pubmed_id = ''
    return project_name,project_title,project_description,project_pubmed_id

def fetch_pubmed(project_pubmed_id: str,iteration: int):
    xml_content = SraUtils.pubmed_id(project_pubmed_id)
    author_list = list()
    grant_list=  list()
    try:
        title = xml_content.find("PubmedArticle").find("MedlineCitation").find("Article").find("ArticleTitle").text
    except:
        title = ''
        if iteration == 1:
            print("no publication title found")
    try:
        authors = xml_content.find("PubmedArticle").find("MedlineCitation").find("Article").find("AuthorList")
    except:
        if iteration == 1:
            print("no authors found in SRA")
        try:
            url = rq.get(f'https://www.ebi.ac.uk/europepmc/webservices/rest/search?query={title}')
            if url.status_code == STATUS_ERROR_CODE:
                raise NotFoundENA(url, title)
            else:
                xml_content_2 = xm.fromstring(url.content)
            try:
                results = list()
                result_list = xml_content_2.find("resultList")
                for result in result_list:
                    results.append(result)
                author_string = results[0].find("authorString").text
            except:
                authors = None
                if iteration == 1:
                    print("no authors found in ENA")
        except:
            authors = None
            if iteration == 1:
                print("no authors found in ENA")
    if authors is not None:
        for author in authors:
            try:
                lastname = author.find("LastName").text
            except:
                lastname = ''
            try:
                forename = author.find("ForeName").text
            except:
                forename = ''
            try:
                initials = author.find("Initials").text
            except:
                initials = ''
            try:
                affiliation = author.find('AffiliationInfo').find("Affiliation").text
            except:
                affiliation = ''
            author_list.append([lastname,forename,initials,affiliation])
    try:
        grants = xml_content.find("PubmedArticle").find("MedlineCitation").find("Article").find("GrantList")
    except:
        grants = None
        if iteration == 1:
            print("no grants found in SRA or ENA")
    if grants:
        for grant in grants:
            try:
                id = grant.find("GrantID").text
            except:
                id = ''
            try:
                agency = grant.find("Agency").text
            except:
                agency = ''
            grant_list.append([id,agency])
    try:
        articles = xml_content.find('PubmedArticle').find('PubmedData').find('ArticleIdList')
        for article_id in articles:
            if "/" in article_id.text:
                article_doi_id = article_id.text
    except:
        article_doi_id = ''
        if iteration == 1:
            print("no publication doi found")
    return title,author_list,grant_list,article_doi_id

def parse_xml(xml_content):
    for experiment_package in xml_content.findall('EXPERIMENT_PACKAGE'):
        yield experiment_package

def get_file_names_from_SRA(experiment_package):
    fastq_map = {}
    run_set = experiment_package.find('RUN_SET')
    for run in run_set:
        try:
            accession = run.attrib['accession']
            run_attributes = run.find('RUN_ATTRIBUTES')
            for attribute in run_attributes:
                if attribute.find('TAG').text == 'options':
                    read_values = attribute.find('VALUE').text
                    fastq_map = extract_read_info(read_values,accession,fastq_map)
        except:
            continue
    return fastq_map

def retrieve_fastq_from_experiment(experiment_packages):
    fastq_list = []
    for experiment_package in experiment_packages:
        fastq_list.extend([get_file_names_from_SRA(element.attrib) for element in experiment_package.iter('SRAFile')])
    return fastq_list

def get_attributes_biosample(element):
    element_id = ''
    if element.attrib:
        element_id = element.attrib['accession']
    if element_id == '':
        for item in element.find('Ids'):
            if 'SAMN' in item.text:
                element_id = item.text
    if element_id == '':
        print('Could not find biosample id')
    sample_title = ''
    for description in element.findall('Description'):
        sample_title = description.find('Title').text
    attribute_list = []
    for attribute_set in element.findall('Attributes'):
        for attribute in attribute_set:
            attribute_list.append(attribute.text)
    if attribute_list == []:
        attribute_list = ['','']
    return [element_id,sample_title,attribute_list]

def get_attributes_library_protocol(experiment_package):
    experiment_id = experiment_package.find('EXPERIMENT').attrib['accession']
    for experiment in experiment_package.find('EXPERIMENT'):
        library_descriptors = experiment.find('LIBRARY_DESCRIPTOR')
        if library_descriptors:
            desc = library_descriptors.find('LIBRARY_CONSTRUCTION_PROTOCOL')
            if desc:
                library_construction_protocol = desc.text
            else:
                library_construction_protocol = ''
        else:
            library_construction_protocol = ''
        illumina = experiment.find('ILLUMINA')
        if illumina:
            instrument = illumina.find('INSTRUMENT_MODEL').text
        else:
            instrument = ''
    return [experiment_id,library_construction_protocol,instrument]

def get_lane_index(file):
    result = re.search('_L[0-9]{3}', file)
    return result

def get_file_index(file):
    if "_I1" in file or "_R3" in file or "_3" in file:
        ind = 'index1'
    elif "_R1" in file or "_1" in file:
        ind = 'read1'
    elif "_R2" in file or "_2" in file:
        ind = 'read2'
    elif "_I2" in file or "_R4" in file or "_4" in file:
        ind = 'index2'
    else:
         ind = ''
    return ind

def integrate_metadata(srp_metadata,fastq_map,cols):
    srp_metadata_update = pd.DataFrame()
    for index, row in srp_metadata.iterrows():
        srr_accession = row['Run']
        if not fastq_map or srr_accession not in fastq_map.keys():
            new_row = row.to_list()
            new_row.extend(['','',''])
            a_series = pd.Series(new_row)
            srp_metadata_update = srp_metadata_update.append(a_series, ignore_index=True)
        else:
            filenames_list = fastq_map[srr_accession]
            for file in filenames_list:
                # Try to find a lane_index. At the moment this is a test to see if they are ever available in
                # fastq file names obtained from ENA. It will not incorporate lane indices if found currently.
                lane_index = get_lane_index(file)
                if lane_index:
                    g = lane_index.group()
                    lane_index = g.split("_")[1]
                else:
                    lane_index = ''
                new_row = row.to_list()
                new_row.extend([file,get_file_index(file),lane_index])
                a_series = pd.Series(new_row)
                srp_metadata_update = srp_metadata_update.append(a_series, ignore_index=True)
    cols.extend(['fastq_name', 'file_index', 'lane_index'])
    srp_metadata_update.columns = cols
    return srp_metadata_update

def get_empty_df(workbook,tab_name):
    sheet = workbook[tab_name]
    values = sheet.values
    empty_df = pd.DataFrame(values)
    cols = empty_df.loc[0,]
    tab = pd.DataFrame(columns=cols)
    return tab

def write_to_wb(workbook: Workbook, tab_name: str, tab_content: pd.DataFrame) -> None:
    """
    Write the tab to the active workbook.
    :param workbook: str
                     Workbook extracted from the template
    :param tab_name: str
                     Name of the tab that is being modified
    :param tab_content: str
                        Content of the tab

    :returns None
    """
    worksheet = workbook[tab_name]
    # If more than 1 series is being filled for the same spreadsheet, find the first unfilled row
    row_not_filled = 6
    while True:
        if worksheet[f'A{row_not_filled}'].value or worksheet[f'B{row_not_filled}'].value:
            row_not_filled += 1
        else:
            break

    for index, key in enumerate(worksheet[4]):
        if not key.value:
            break
        if key.value not in tab_content.keys():
            continue
        for i in range(len(tab_content[key.value])):
            worksheet[f"{get_column_letter(index + 1)}{i + row_not_filled}"] = list(tab_content[key.value])[i]

def get_sequence_file_tab_xls(SRP_df,workbook,tab_name):
    tab = get_empty_df(workbook,tab_name)
    for index,row in SRP_df.iterrows():
        tab = tab.append({'sequence_file.file_core.file_name': row['fastq_name'],
                          'sequence_file.file_core.format': 'fastq.gz',
                          'sequence_file.file_core.content_description.text':'DNA sequence',
                          'sequence_file.read_index': row['file_index'],
                          'sequence_file.lane_index': row['lane_index'],
                          'sequence_file.insdc_run_accessions': row['Run'],
                          'process.insdc_experiment.insdc_experiment_accession': row['Experiment'],
                          'cell_suspension.biomaterial_core.biomaterial_id':row['Experiment'],
                          'library_preparation_protocol.protocol_core.protocol_id':'',
                          'sequencing_protocol.protocol_core.protocol_id':'',
                          'process.process_core.process_id':row['Run']}, ignore_index=True)
    tab = tab.sort_values(by='sequence_file.insdc_run_accessions')
    return tab

def get_cell_suspension_tab_xls(SRP_df,workbook,out_file,tab_name):
    tab = get_empty_df(workbook, tab_name)
    experiments_dedup = list(set(list(SRP_df['Experiment'])))
    for experiment in experiments_dedup:
        biosample = list(SRP_df.loc[SRP_df['Experiment'] == experiment]['BioSample'])[0]
        gsm_sample = list(SRP_df.loc[SRP_df['Experiment'] == experiment]['SampleName'])[0]
        tab = tab.append({'cell_suspension.biomaterial_core.biomaterial_id':experiment,
                          'cell_suspension.biomaterial_core.biomaterial_name':gsm_sample,
                         'specimen_from_organism.biomaterial_core.biomaterial_id':biosample,
                          'cell_suspension.biomaterial_core.ncbi_taxon_id':list(SRP_df.loc[SRP_df['Experiment'] == experiment]['TaxID'])[0],
                         'cell_suspension.genus_species.text':list(SRP_df.loc[SRP_df['Experiment'] == experiment]['ScientificName'])[0],
                         'cell_suspension.biomaterial_core.biosamples_accession':biosample}, ignore_index=True)
    tab = tab.sort_values(by='cell_suspension.biomaterial_core.biomaterial_id')
    write_to_wb(workbook, tab_name, tab)

def process_specimen_from_organism(biosample_attribute_list,srp_metadata_update):
    df = {'specimen_from_organism.biomaterial_core.biomaterial_id':biosample_attribute_list[0],
          'specimen_from_organism.biomaterial_core.biomaterial_name':biosample_attribute_list[1],
          'specimen_from_organism.biomaterial_core.biomaterial_description': ','.join(biosample_attribute_list[2]),
          'specimen_from_organism.biomaterial_core.ncbi_taxon_id': list(srp_metadata_update.loc[srp_metadata_update['BioSample'] == biosample_attribute_list[0]]['TaxID'])[0],
          'specimen_from_organism.genus_species.text': list(srp_metadata_update.loc[srp_metadata_update['BioSample'] == biosample_attribute_list[0]]['ScientificName'])[0],
          'specimen_from_organism.genus_species.ontology_label': list(srp_metadata_update.loc[srp_metadata_update['BioSample'] == biosample_attribute_list[0]]['ScientificName'])[0],
          'specimen_from_organism.biomaterial_core.biosamples_accession': biosample_attribute_list[0],
          'specimen_from_organism.biomaterial_core.insdc_sample_accession': list(srp_metadata_update.loc[srp_metadata_update['BioSample'] == biosample_attribute_list[0]]['Sample'])[0],
          'collection_protocol.protocol_core.protocol_id':'',
          'process.insdc_experiment.insdc_experiment_accession':srp_metadata_update[srp_metadata_update['BioSample'] == biosample_attribute_list[0]]['Experiment'].values.tolist()[0]}
    return df

@contextmanager
def poolcontext(*args, **kwargs):
    pool = multiprocessing.Pool(*args, **kwargs)
    yield pool
    pool.terminate()

def get_specimen_from_organism_tab_xls(srp_metadata_update,workbook,nthreads,tab_name):
    tab = get_empty_df(workbook, tab_name)
    biosample_accessions = list(set(list(srp_metadata_update['BioSample'])))
    attribute_lists = fetch_accession_info(biosample_accessions,accession_type='biosample')
    results = None
    if attribute_lists:
        try:
            with poolcontext(processes=nthreads) as pool:
                results = pool.map(partial(process_specimen_from_organism, srp_metadata_update=srp_metadata_update), attribute_lists)
        except KeyboardInterrupt:
            print("Process has been interrupted.")
            pool.terminate()
    if results:
        df = pd.DataFrame(results)
        tab = tab.append(df,sort=True)
        tab = tab.sort_values(by='process.insdc_experiment.insdc_experiment_accession')
        write_to_wb(workbook, tab_name, tab)

def get_library_protocol_tab_xls(SRP_df,workbook,tab_name):
    tab = get_empty_df(workbook, tab_name)
    experiment_accessions = list(set(list(SRP_df['Experiment'])))
    count = 0
    library_protocol_set = list()
    library_protocol_dict = {}
    attribute_lists = fetch_accession_info(experiment_accessions,accession_type='experiment')
    for attribute_list in attribute_lists:
        experiment_accession = str(attribute_list[0])
        library_protocol = attribute_list[1]
        if library_protocol not in library_protocol_set:
            count += 1
            library_protocol_id = "library_protocol_" + str(count)
            library_protocol_set.append(library_protocol)
            tmp_dict = {'library_preparation_protocol.protocol_core.protocol_id':library_protocol_id,
                              'library_preparation_protocol.protocol_core.protocol_description': library_protocol,
                              'library_preparation_protocol.input_nucleic_acid_molecule.text': 'polyA RNA',
                              'library_preparation_protocol.nucleic_acid_source':'single cell'}
            library_protocol_dict[experiment_accession] = {"library_protocol_id":library_protocol_id,"library_protocol_description":library_protocol}
            if "10X" in library_protocol:
                if "v.2" or "v2" or 'V2' or 'V.2' in library_protocol:
                    if "3'" in library_protocol:
                        tmp_dict.update({'library_preparation_protocol.cell_barcode.barcode_read': 'Read1',
                                        'library_preparation_protocol.cell_barcode.barcode_offset': 0,
                                        'library_preparation_protocol.cell_barcode.barcode_length': 16,
                                        'library_preparation_protocol.library_construction_method.text':"10X 3' v2 sequencing",
                                        'library_preparation_protocol.library_construction_kit.retail_name': 'Single Cell 3’ Reagent Kit v2',
                                        'library_preparation_protocol.library_construction_kit.manufacturer': '10X Genomics',
                                        'library_preparation_protocol.end_bias':'3 prime tag',
                                        'library_preparation_protocol.primer':'poly-dT',
                                        'library_preparation_protocol.strand':'first',
                                        'library_preparation_protocol.umi_barcode.barcode_read':'Read1',
                                        'library_preparation_protocol.umi_barcode.barcode_offset':16,
                                        'library_preparation_protocol.umi_barcode.barcode_length':10})
                    elif "5'" in library_protocol:
                        print("Please let Ami know that you have come across a 10X v2 5' dataset")
                        tmp_dict.update({'library_preparation_protocol.cell_barcode.barcode_read': 'Read1',
                                        'library_preparation_protocol.cell_barcode.barcode_offset': 0,
                                        'library_preparation_protocol.cell_barcode.barcode_length': 16,
                                        'library_preparation_protocol.library_construction_method.text':"10X 5' v2 sequencing",
                                        'library_preparation_protocol.library_construction_kit.retail_name': 'Single Cell 5’ Reagent Kit v2',
                                        'library_preparation_protocol.library_construction_kit.manufacturer': '10X Genomics',
                                        'library_preparation_protocol.end_bias':'5 prime tag',
                                        'library_preparation_protocol.primer':'poly-dT',
                                        'library_preparation_protocol.strand':'first',
                                        'library_preparation_protocol.umi_barcode.barcode_read':'Read1',
                                        'library_preparation_protocol.umi_barcode.barcode_offset':16,
                                        'library_preparation_protocol.umi_barcode.barcode_length':10})
                elif "v.3" or "v3" or 'V3' or 'V.3' in library_protocol:
                    tmp_dict.update({'library_preparation_protocol.cell_barcode.barcode_read': 'Read1',
                                    'library_preparation_protocol.cell_barcode.barcode_offset': 0,
                                    'library_preparation_protocol.cell_barcode.barcode_length': 16,
                                    'library_preparation_protocol.library_construction_method.text':"10X 3' v3 sequencing",
                                    'library_preparation_protocol.library_construction_kit.retail_name': 'Single Cell 3’ Reagent Kit v3',
                                    'library_preparation_protocol.library_construction_kit.manufacturer': '10X Genomics',
                                    'library_preparation_protocol.end_bias':'3 prime tag',
                                    'library_preparation_protocol.primer':'poly-dT',
                                    'library_preparation_protocol.strand':'first',
                                    'library_preparation_protocol.umi_barcode.barcode_read':'"Read1',
                                    'library_preparation_protocol.umi_barcode.barcode_offset':16,
                                    'library_preparation_protocol.umi_barcode.barcode_length':12})
                elif "v.1" or "v1" or 'V1' or 'V.1' in library_protocol:
                    tmp_dict.update({'library_preparation_protocol.cell_barcode.barcode_read': 'Read1',
                                    'library_preparation_protocol.cell_barcode.barcode_offset': 0,
                                    'library_preparation_protocol.cell_barcode.barcode_length': 14,
                                    'library_preparation_protocol.library_construction_method.text':"10X v1 sequencing",
                                    'library_preparation_protocol.library_construction_kit.retail_name': 'Single Cell Reagent Kit v1',
                                    'library_preparation_protocol.library_construction_kit.manufacturer': '10X Genomics',
                                    'library_preparation_protocol.end_bias':'',
                                    'library_preparation_protocol.primer':'poly-dT',
                                    'library_preparation_protocol.strand':'first',
                                    'library_preparation_protocol.umi_barcode.barcode_read':'Read1',
                                    'library_preparation_protocol.umi_barcode.barcode_offset':14,
                                    'library_preparation_protocol.umi_barcode.barcode_length':10})
                else:
                    tmp_dict.update({'library_preparation_protocol.library_construction_method.text': "10X sequencing",
                                     'library_preparation_protocol.library_construction_kit.manufacturer': '10X Genomics'})
                tab = tab.append(tmp_dict, ignore_index=True)
            elif 'Drop-seq' or 'drop-seq' or 'DropSeq' or 'Dropseq' in library_protocol:
                tmp_dict.update({'library_preparation_protocol.cell_barcode.barcode_read': 'Read1',
                                 'library_preparation_protocol.cell_barcode.barcode_offset': 0,
                                 'library_preparation_protocol.cell_barcode.barcode_length': 12,
                                 'library_preparation_protocol.library_construction_method.text': "Drop-seq",
                                 'library_preparation_protocol.library_construction_kit.retail_name': '',
                                 'library_preparation_protocol.library_construction_kit.manufacturer': '',
                                 'library_preparation_protocol.end_bias': '',
                                 'library_preparation_protocol.primer': 'poly-dT',
                                 'library_preparation_protocol.strand': 'first',
                                 'library_preparation_protocol.umi_barcode.barcode_read': 'Read1',
                                 'library_preparation_protocol.umi_barcode.barcode_offset': 12,
                                 'library_preparation_protocol.umi_barcode.barcode_length': 8})
                tab = tab.append(tmp_dict, ignore_index=True)
            elif 'Smart-seq' or 'smart-seq' or 'Smartseq' or 'SmartSeq' or 'plate' or 'Plate' in library_protocol:
                tmp_dict.update({'library_preparation_protocol.cell_barcode.barcode_read': '',
                                 'library_preparation_protocol.cell_barcode.barcode_offset': '',
                                 'library_preparation_protocol.cell_barcode.barcode_length': '',
                                 'library_preparation_protocol.library_construction_method.text': 'Smart-seq2',
                                 'library_preparation_protocol.library_construction_kit.retail_name': '',
                                 'library_preparation_protocol.library_construction_kit.manufacturer': '',
                                 'library_preparation_protocol.end_bias': 'full length',
                                 'library_preparation_protocol.primer': 'poly-dT',
                                 'library_preparation_protocol.strand': 'unstranded',
                                 'library_preparation_protocol.umi_barcode.barcode_read': '',
                                 'library_preparation_protocol.umi_barcode.barcode_offset': '',
                                 'library_preparation_protocol.umi_barcode.barcode_length': ''})
                tab = tab.append(tmp_dict, ignore_index=True)
            else:
                tab = tab
        if library_protocol in library_protocol_set:
            tab = tab
            for key in library_protocol_dict.keys():
                if library_protocol_dict[key]["library_protocol_description"] == library_protocol:
                    library_protocol_id = library_protocol_dict[key]["library_protocol_id"]
            if not library_protocol_id:
                library_protocol_id = ''
            else:
                library_protocol_dict[experiment_accession] = {"library_protocol_id":library_protocol_id,"library_protocol_description":
                library_protocol}
    write_to_wb(workbook, tab_name, tab)
    return library_protocol_dict,attribute_lists

def get_sequencing_protocol_tab_xls(SRP_df,workbook,attribute_lists,tab_name):
    tab = get_empty_df(workbook, tab_name)
    count = 0
    sequencing_protocol_id = "sequencing_protocol_1"
    sequencing_protocol_set = list()
    sequencing_protocol_dict = {}
    for attribute_list in attribute_lists:
        experiment = attribute_list[0]
        library_construction_protocol = attribute_list[1]
        instrument = attribute_list[2]
        if "10X" in library_construction_protocol:
            paired_end = 'no'
            method = 'tag based single cell RNA sequencing'
        elif "10X" not in library_construction_protocol:
            paired_end = ''
            method = ''
        sequencing_protocol_description = [instrument,method]
        if sequencing_protocol_description not in sequencing_protocol_set:
            count += 1
            sequencing_protocol_id = "sequencing_protocol_" + str(count)
            sequencing_protocol_set.append(sequencing_protocol_description)
            sequencing_protocol_dict[experiment] = {"sequencing_protocol_id":sequencing_protocol_id,
                                                    "sequencing_protocol_description":sequencing_protocol_description}
            tab = tab.append({'sequencing_protocol.protocol_core.protocol_id': sequencing_protocol_id,
                              'sequencing_protocol.instrument_manufacturer_model.text': instrument,
                              'sequencing_protocol.paired_end': paired_end,
                              'sequencing_protocol.method.text': method}, ignore_index=True)
        if sequencing_protocol_description in sequencing_protocol_set:
            tab = tab
            for key in sequencing_protocol_dict.keys():
                if sequencing_protocol_dict[key]["sequencing_protocol_description"] == sequencing_protocol_description:
                    sequencing_protocol_id = sequencing_protocol_dict[key]["sequencing_protocol_id"]
            if not sequencing_protocol_id:
                sequencing_protocol_id = ''
            else:
                sequencing_protocol_dict[experiment] = {"sequencing_protocol_id":sequencing_protocol_id,
                                                    "sequencing_protocol_description":sequencing_protocol_description}
    write_to_wb(workbook, tab_name, tab)
    return sequencing_protocol_dict

def update_sequence_file_tab_xls(sequence_file_tab,library_protocol_dict,sequencing_protocol_dict,workbook,tab_name):
    library_protocol_id_list = list()
    sequencing_protocol_id_list = list()
    for index,row in sequence_file_tab.iterrows():
        if row["cell_suspension.biomaterial_core.biomaterial_id"] in library_protocol_dict.keys():
            library_protocol_id_list.append(library_protocol_dict[row["cell_suspension.biomaterial_core.biomaterial_id"]]["library_protocol_id"])
        else:
            library_protocol_id_list.append('')
        if row["cell_suspension.biomaterial_core.biomaterial_id"] in sequencing_protocol_dict.keys():
            sequencing_protocol_id_list.append(sequencing_protocol_dict[row["cell_suspension.biomaterial_core.biomaterial_id"]]["sequencing_protocol_id"])
        else:
            sequencing_protocol_id_list.append('')
    sequence_file_tab['library_preparation_protocol.protocol_core.protocol_id'] = library_protocol_id_list
    sequence_file_tab['sequencing_protocol.protocol_core.protocol_id'] = sequencing_protocol_id_list
    write_to_wb(workbook, tab_name, sequence_file_tab)

def get_project_main_tab_xls(SRP_df,workbook,geo_accession,out_file,tab_name):
    study = list(SRP_df['SRAStudy'])[0]
    project = list(SRP_df['BioProject'])[0]
    try:
        tab = get_empty_df(workbook,tab_name)
        bioproject = list(set(list(SRP_df['BioProject'])))
        if len(bioproject) > 1:
            print("more than 1 bioproject, check this")
        else:
            bioproject = bioproject[0]
        project_name,project_title,project_description,project_pubmed_id = fetch_bioproject(bioproject)
        tab = tab.append({'project.project_core.project_title':project_title,
                          'project.project_core.project_description':project_description,
                          'project.geo_series_accessions':geo_accession,
                          'project.insdc_study_accessions':study,
                          'project.insdc_project_accessions':project}, ignore_index=True)
        write_to_wb(workbook, tab_name, tab)
    except AttributeError:
        pass
    return project_name,project_title,project_description,project_pubmed_id


def get_project_publication_tab_xls(workbook,tab_name,project_pubmed_id):
    tab = get_empty_df(workbook,tab_name)
    title,author_list,grant_list,article_doi_id = fetch_pubmed(project_pubmed_id,iteration=1)
    name_list = list()
    for author in author_list:
        name = author[0] + ' ' + author[2] + "||"
        name_list.append(name)
    name_list = ''.join(name_list)
    name_list = name_list[:len(name_list)-2]
    tab = tab.append({'project.publications.authors':name_list,
                      'project.publications.title':title,
                      'project.publications.doi':article_doi_id,
                      'project.publications.pmid':project_pubmed_id,
                      'project.publications.url':''}, ignore_index=True)
    write_to_wb(workbook, tab_name, tab)


def get_project_contributors_tab_xls(workbook,tab_name,project_pubmed_id):
    tab = get_empty_df(workbook,tab_name)
    title, author_list, grant_list, article_doi_id = fetch_pubmed(project_pubmed_id,iteration=2)
    for author in author_list:
        name = author[1] + ',,' + list(author)[0]
        affiliation = author[3]
        tab = tab.append({'project.contributors.name':name,'project.contributors.institution':affiliation}, ignore_index=True)
    write_to_wb(workbook, tab_name, tab)


def get_project_funders_tab_xls(workbook,tab_name,project_pubmed_id):
    tab = get_empty_df(workbook,tab_name)
    title, author_list, grant_list, article_doi_id = fetch_pubmed(project_pubmed_id,iteration=3)
    for grant in grant_list:
        tab = tab.append({'project.funders.grant_id':grant[0],'project.funders.organization':grant[1]}, ignore_index=True)
    write_to_wb(workbook, tab_name, tab)


def empty_worksheet(worksheet):
    if worksheet['A6'].value or worksheet['B6'].value:
        return False
    return True

def delete_unused_worksheets(workbook: Workbook) -> None:
    """
    Delete unused sheets from the metadata spreadsheet and the linked protocols.

    :param workbook: Workbook
                     Workbook object containing the metadata spreadsheet being modified
    :returns None
    """

    for worksheet_name in OPTIONAL_TABS:
        current_worksheet = workbook[worksheet_name]
        if current_worksheet and empty_worksheet(current_worksheet):
            del workbook[worksheet_name]
            if worksheet_name in LINKINGS:
                for linked_sheet in LINKINGS[worksheet_name]:
                    if empty_worksheet(workbook[linked_sheet]):
                        del workbook[linked_sheet]

def list_str(values):
    if "," not in values:
        raise argparse.ArgumentTypeError("Argument list not valid: comma separated list required")
    return values.split(',')

def check_file(path):
    if not os.path.exists(path):
        raise argparse.ArgumentTypeError("file %s does not exist" % (path))
    try:
        df = pd.read_csv(path, sep="\t")
    except:
        raise argparse.ArgumentTypeError("file %s is not a valid format" % (path))
    try:
        geo_accession_list = list(df["accession"])
    except:
        raise argparse.ArgumentTypeError("accession list column not found in file %s" % (path))
    return geo_accession_list

def main():

    parser = argparse.ArgumentParser()
    parser.add_argument('--accession',type=str,help='accession (str): either GEO or SRA accession')
    parser.add_argument('--accession_list',type=list_str,help='accession list (comma separated)')
    parser.add_argument('--input_file',type=check_file,help='optional path to tab-delimited input .txt file')
    parser.add_argument('--nthreads',type=int,default=1,
                        help='number of multiprocessing processes to use')
    parser.add_argument('--template',default="docs/hca_template.xlsx",
                        help='path to an HCA spreadsheet template (xlsx)')
    parser.add_argument('--header_row',type=int,default=4,
                        help='header row with HCA programmatic names')
    parser.add_argument('--input_row1',type=int,default=6,
                        help='HCA metadata input start row')
    parser.add_argument('--output_dir',default='spreadsheets/',
                        help='path to output directory; if it does not exist, the directory will be created')
    parser.add_argument('--output_log',type=bool,default=True,
                        help='True/False: should the output result log be created')

    args = parser.parse_args()

    # check user-specified arguments are valid

    if args.input_file:
        accession_list = args.input_file
    elif args.accession_list:
        accession_list = args.accession_list
    elif args.accession:
        accession_list = [args.accession]
    else:
        print("GEO or SRA accession input is not specified")
        sys.exit()

    if not os.path.exists(args.template):
        print("path to HCA template file not found; will revert to default: docs/hca_template.xlsx")
        template = "docs/hca_template.xlsx"
    try:
        workbook = load_workbook(filename=args.template)
        template = args.template
    except:
        print("specified HCA template file is not valid xlsx; will revert to default: docs/hca_template.xlsx")
        template = "docs/hca_template.xlsx"

    if not os.path.exists(args.output_dir):
        os.mkdir(args.output_dir)

    # initialise dictionary to summarise results
    results = {}

    # for each geo accession:
    for accession in accession_list:

        # create a new output file name to store the hca converted metadata
        out_file = f"{args.output_dir}/{accession}.xlsx"

        # load an empty template HCA metadata excel spreadsheet. All tabs and fields should be in this template.
        workbook = load_workbook(filename=template)

        srp_accession = None

        if 'GSE' in accession:

            print(f"Fetching SRA study ID for GEO dataset {accession}")
            # fetch the SRA study accession given the geo accession
            srp_accession = fetch_srp_accession(accession)
            print(f"Found SRA study ID: {srp_accession}")

        elif 'SRP' in accession:
            srp_accession = accession

        if not srp_accession:
            results[accession] = {"SRA Study available": "no"}
            results[accession].update({"fastq files available": "na"})
            print(f"No SRA study accession is available for accession {accession}")
            continue

        else:
            # fetch the SRA study metadata for the srp accession
            print(f"Fetching study metadata for SRA study ID: {srp_accession}")
            srp_metadata = fetch_srp_metadata(srp_accession)

            # Get dataframe column names for later
            cols = srp_metadata.columns.tolist()

            # get fastq file names
            print(f"Fetching fastq file names for SRA study ID: {srp_accession}")
            fastq_map = fetch_fastq_names(srp_accession, list(srp_metadata['Run']))

            if not fastq_map:

                print(f"Both Read1 and Read2 fastq files are not available for SRA study ID: {srp_accession}")
                results[accession] = {"SRA Study available": "yes"}
                results[accession].update({"fastq files available": "no"})

            else:

                print(f"Found fastq files for SRA study ID: {srp_accession}")
                results[accession] = {"SRA Study available": "yes"}
                results[accession].update({"fastq files available": "yes"})

            # integrate metadata and fastq file names into a single dataframe
            print(f"Integrating study metadata and fastq file names")
            srp_metadata_update = integrate_metadata(srp_metadata, fastq_map, cols)

            print(f"Getting Sequence file tab")
            # get HCA Sequence file metadata: fetch as many fields as is possible using the above metadata accessions
            sequence_file_tab = get_sequence_file_tab_xls(srp_metadata_update,workbook,tab_name="Sequence file")

            print(f"Getting Cell suspension tab")
            # get HCA Cell suspension metadata: fetch as many fields as is possible using the above metadata accessions
            get_cell_suspension_tab_xls(srp_metadata_update,workbook,out_file,tab_name="Cell suspension")

            print(f"Getting Specimen from Organism tab")
            # get HCA Specimen from organism metadata: fetch as many fields as is possible using the above metadata accessions
            get_specimen_from_organism_tab_xls(srp_metadata_update,workbook,args.nthreads,tab_name="Specimen from organism")

            print(f"Getting Library preparation protocol tab")
            # get HCA Library preparation protocol metadata: fetch as many fields as is possible using the above metadata accessions
            library_protocol_dict,attribute_lists = get_library_protocol_tab_xls(srp_metadata_update,workbook,tab_name="Library preparation protocol")

            print(f"Getting Sequencing protocol tab")
            # get HCA Sequencing protocol metadata: fetch as many fields as is possible using the above metadata accessions
            sequencing_protocol_dict = get_sequencing_protocol_tab_xls(srp_metadata_update,workbook,attribute_lists,tab_name="Sequencing protocol")

            print(f"Updating Sequencing file tab with protocol ids")
            # update HCA Sequence file metadata with the correct library preparation protocol ids and sequencing protocol ids
            update_sequence_file_tab_xls(sequence_file_tab,library_protocol_dict,sequencing_protocol_dict,workbook,tab_name="Sequence file")

            print(f"Getting project metadata")
            # get Project metadata: fetch as many fields as is possible using the above metadata accessions
            project_name, project_title, project_description, project_pubmed_id = get_project_main_tab_xls(srp_metadata_update,workbook,accession,out_file,tab_name="Project")

            try:
                # get Project - Publications metadata: fetch as many fields as is possible using the above metadata accessions
                get_project_publication_tab_xls(workbook,tab_name="Project - Publications",project_pubmed_id=project_pubmed_id)
            except AttributeError:
                print(f'Publication attribute error with GEO project {accession}')

            try:
                # get Project - Contributors metadata: fetch as many fields as is possible using the above metadata accessions
                get_project_contributors_tab_xls(workbook,tab_name="Project - Contributors",project_pubmed_id=project_pubmed_id)
            except AttributeError:
                print(f'Contributors attribute error with GEO project {accession}')

            try:
                # get Project - Funders metadata: fetch as many fields as is possible using the above metadata accessions
                get_project_funders_tab_xls(workbook,tab_name="Project - Funders",project_pubmed_id=project_pubmed_id)
            except AttributeError:
                print(f'Funders attribute error with GEO project {accession}')

        # Done
        print(f"Done. Saving workbook to excel file")
        workbook.save(out_file)

    results = pd.DataFrame.from_dict(results).transpose()
    print("showing result")
    print(results)
    if args.output_log:
        results.to_csv(f"{args.output_dir}/results_{accession}.log",sep="\t")
    print("Done.")


if __name__ == "__main__":
    main()
